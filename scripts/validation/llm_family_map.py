#!/usr/bin/env python3
"""
LLM-assisted posting-to-family mapper for the validation study.

Offline validation tooling only. The mapper classifies a LinkedIn posting into
one of the 21 retained resume families plus UNMAPPED. The scoped validation then
filters those cached labels down to INFORMATION-TECHNOLOGY, FINANCE, and
CONSULTANT; the mapper itself intentionally remains 22-way so future human
checks can benchmark the full mapping task.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import time
import urllib.error
import urllib.request
from copy import copy
from pathlib import Path
from typing import Any


HERE = Path(__file__).resolve().parent
REPO = HERE.parent.parent
ART = HERE / ".artifacts"
ENV_PATH = REPO / ".env.local"

DEFAULT_MODEL = "gpt-4o-mini"
OPENAI_URL = "https://api.openai.com/v1/chat/completions"

ALLOWED_FAMILIES = [
    "ACCOUNTANT",
    "ADVOCATE",
    "AVIATION",
    "BANKING",
    "CHEF",
    "CONSTRUCTION",
    "CONSULTANT",
    "PUBLIC-RELATIONS",
    "DIGITAL-MEDIA",
    "HR",
    "FINANCE",
    "BUSINESS-DEVELOPMENT",
    "SALES",
    "HEALTHCARE",
    "FITNESS",
    "TEACHER",
    "DESIGNER",
    "ARTS",
    "APPAREL",
    "INFORMATION-TECHNOLOGY",
    "ENGINEERING",
    "UNMAPPED",
]

FAMILY_DEFINITIONS = {
    "ACCOUNTANT": "Accounting, bookkeeping, audit, tax, controller, accounts payable/receivable, and accounting operations roles.",
    "ADVOCATE": "Legal roles such as attorney, lawyer, paralegal, legal counsel, litigation, law clerk, and legal assistant roles.",
    "AVIATION": "Aviation and aerospace roles such as pilot, aircraft, avionics, airline operations, flight attendant, and aircraft maintenance roles.",
    "BANKING": "Banking roles such as teller, loan officer, mortgage, credit analyst, branch banking, underwriting, and personal banker roles.",
    "CHEF": "Culinary and food preparation roles such as chef, cook, kitchen, bakery, pastry, sous chef, and restaurant kitchen roles.",
    "CONSTRUCTION": "Construction, civil engineering, contractor, site supervisor, surveyor, carpenter, estimator, and field construction roles.",
    "CONSULTANT": "Management consulting, business consulting, advisory, process improvement, transformation, and client-facing analysis roles.",
    "PUBLIC-RELATIONS": "Public relations, media relations, communications, press, publicity, publicist, and external communications roles.",
    "DIGITAL-MEDIA": "Digital marketing, social media, SEO/SEM, content strategy, online media, copywriting, and community management roles.",
    "HR": "Human resources, recruiting, talent acquisition, people operations, benefits, HRIS, employee relations, and HR business partner roles.",
    "FINANCE": "Financial analysis, investment, FP&A, treasury, portfolio, valuation, equity research, and corporate finance roles.",
    "BUSINESS-DEVELOPMENT": "Business development, partnerships, strategic accounts, growth partnerships, and business developer roles.",
    "SALES": "Sales roles such as account executive, account manager, sales representative, retail sales, inside sales, and sales management roles.",
    "HEALTHCARE": "Clinical and healthcare delivery roles such as nurse, physician, therapist, pharmacist, medical assistant, caregiver, and patient care roles.",
    "FITNESS": "Fitness, personal training, gym, yoga instructor, wellness coach, athletic training, and group fitness roles.",
    "TEACHER": "Education roles such as teacher, professor, instructor, tutor, lecturer, educator, faculty, and teaching assistant roles.",
    "DESIGNER": "UX, UI, graphic, product, web, interaction, and visual design roles.",
    "ARTS": "Arts and creative roles such as artist, photographer, musician, illustrator, animator, painter, sculptor, and art director roles.",
    "APPAREL": "Fashion, apparel, textile, merchandiser, stylist, tailor, garment, and fashion design roles.",
    "INFORMATION-TECHNOLOGY": "Software, data, cloud, cybersecurity, database, systems administration, IT support, DevOps, QA, and web development roles.",
    "ENGINEERING": "Engineering roles such as mechanical, electrical, manufacturing, process, quality, industrial, chemical, and non-software engineering roles.",
    "UNMAPPED": "Use when the posting does not clearly fit any allowed resume family.",
}

KAGGLE_HEADERS = [
    "job_id",
    "company_name",
    "title",
    "description",
    "max_salary",
    "pay_period",
    "location",
    "company_id",
    "views",
    "med_salary",
    "min_salary",
    "formatted_work_type",
    "applies",
    "original_listed_time",
    "remote_allowed",
    "job_posting_url",
    "application_url",
    "application_type",
    "expiry",
    "closed_time",
    "formatted_experience_level",
    "skills_desc",
    "listed_time",
    "posting_domain",
    "sponsored",
    "work_type",
    "currency",
    "compensation_type",
    "normalized_salary",
    "zip_code",
    "fips",
]

OUTPUT_COLUMNS = ["llm_family", "llm_confidence", "llm_rationale", "llm_error"]


def load_dotenv(path: Path = ENV_PATH) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def clean_cell(value: Any, max_chars: int = 4500) -> str:
    if value is None:
        return ""
    text = str(value).replace("\x00", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text[:max_chars]


def posting_cache_key(posting: dict[str, str]) -> str:
    job_id = clean_cell(posting.get("job_id", ""), max_chars=120)
    if job_id:
        return f"job:{job_id}"
    sig = "\0".join(
        clean_cell(posting.get(k, ""), max_chars=10000)
        for k in ("title", "company_name", "description")
    )
    return "sha1:" + hashlib.sha1(sig.encode("utf-8")).hexdigest()


def prompt_for_posting(posting: dict[str, str]) -> str:
    definitions = "\n".join(
        f"- {family}: {FAMILY_DEFINITIONS[family]}" for family in ALLOWED_FAMILIES
    )
    return f"""
Classify the LinkedIn job posting into exactly one allowed resume family.

Rules:
- Use the title, description, and skills_desc. Do not rely on title alone.
- Classify by the primary role/function, not just the employer's industry.
- Choose UNMAPPED if none of the families clearly fit.
- Do not invent a new family.
- Return JSON only with keys: family, confidence, rationale.
- confidence must be a number from 0 to 1.
- family must be one of the allowed families.

Allowed families and definitions:
{definitions}

Posting:
company_name: {posting.get("company_name", "")}
title: {posting.get("title", "")}
description: {clean_cell(posting.get("description", ""), max_chars=3500)}
skills_desc: {clean_cell(posting.get("skills_desc", ""), max_chars=1000)}
""".strip()


def parse_json_object(text: str) -> dict[str, Any]:
    fenced = re.search(r"```(?:json)?\s*([\s\S]*?)```", text, re.I)
    raw = fenced.group(1) if fenced else text
    start = raw.find("{")
    end = raw.rfind("}")
    if start == -1 or end == -1 or end < start:
        raise ValueError("invalid_json")
    return json.loads(raw[start : end + 1])


def call_openai(prompt: str, api_key: str, model: str) -> dict[str, Any]:
    payload = {
        "model": model,
        "temperature": 0,
        "response_format": {"type": "json_object"},
        "messages": [
            {
                "role": "system",
                "content": "You are a careful job-posting classifier. Return only valid JSON.",
            },
            {"role": "user", "content": prompt},
        ],
    }
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        OPENAI_URL,
        data=data,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as res:
            body = json.loads(res.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"http_{exc.code}: {detail[:500]}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"request_failed: {exc.reason}") from exc

    content = body["choices"][0]["message"]["content"]
    out = parse_json_object(content)
    family = str(out.get("family", "")).strip().upper()
    if family not in ALLOWED_FAMILIES:
        raise ValueError(f"invalid_family: {family}")

    try:
        confidence: float | str = max(0.0, min(1.0, float(out.get("confidence", ""))))
    except (TypeError, ValueError):
        confidence = ""

    return {
        "family": family,
        "confidence": confidence,
        "rationale": clean_cell(out.get("rationale", ""), max_chars=1000),
    }


def classify_posting(
    posting: dict[str, str],
    api_key: str,
    model: str = DEFAULT_MODEL,
) -> dict[str, Any]:
    result = call_openai(prompt_for_posting(posting), api_key=api_key, model=model)
    return {
        "cache_key": posting_cache_key(posting),
        "job_id": clean_cell(posting.get("job_id", ""), max_chars=120),
        "title": clean_cell(posting.get("title", ""), max_chars=300),
        "company_name": clean_cell(posting.get("company_name", ""), max_chars=200),
        "family": result["family"],
        "confidence": result["confidence"],
        "rationale": result["rationale"],
        "error": None,
        "model": model,
    }


def read_jsonl_cache(path: Path) -> dict[str, dict[str, Any]]:
    cache: dict[str, dict[str, Any]] = {}
    if not path.exists():
        return cache
    with path.open("r", encoding="utf-8") as fh:
        for line in fh:
            if not line.strip():
                continue
            row = json.loads(line)
            key = row.get("cache_key")
            if key:
                cache[str(key)] = row
    return cache


def append_jsonl(path: Path, row: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(row, ensure_ascii=False) + "\n")


# --------------------------------------------------------------------------- #
# Optional workbook mode retained for manual samples.
# --------------------------------------------------------------------------- #
def _load_workbook():
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - local environment guard
        raise SystemExit("Missing dependency: openpyxl. Install it to process .xlsx files.") from exc
    return load_workbook


def _looks_like_header(values: list[Any]) -> bool:
    lowered = {str(v).strip().lower() for v in values if v is not None}
    return "title" in lowered and "description" in lowered


def _ensure_headers(ws) -> dict[str, int]:
    first = [cell.value for cell in ws[1]]
    if not _looks_like_header(first):
        ws.insert_rows(1)
        for i in range(1, ws.max_column + 1):
            header = KAGGLE_HEADERS[i - 1] if i <= len(KAGGLE_HEADERS) else f"col_{i}"
            ws.cell(row=1, column=i, value=header)

    headers: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[str(cell.value).strip().lower()] = cell.column

    for name in OUTPUT_COLUMNS:
        if name not in headers:
            col = ws.max_column + 1
            ws.cell(row=1, column=col, value=name)
            headers[name] = col

    missing = [name for name in ("title", "description") if name not in headers]
    if missing:
        raise SystemExit(f"Missing required column(s): {', '.join(missing)}")
    return headers


def _copy_cell_style(src, dst) -> None:
    if not src.has_style:
        return
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def _style_output_headers(ws, headers: dict[str, int]) -> None:
    base = ws.cell(row=1, column=max(1, headers.get("description", 1)))
    for name in OUTPUT_COLUMNS:
        _copy_cell_style(base, ws.cell(row=1, column=headers[name]))


def _row_text(ws, row: int, headers: dict[str, int], name: str) -> str:
    col = headers.get(name)
    return clean_cell(ws.cell(row=row, column=col).value) if col else ""


def classify_workbook(args: argparse.Namespace) -> None:
    load_dotenv()
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    model = os.environ.get("OPENAI_MODEL", DEFAULT_MODEL).strip() or DEFAULT_MODEL
    if not api_key and not args.dry_run:
        raise SystemExit("Missing OPENAI_API_KEY. Add it to .env.local or the environment.")

    input_path = Path(args.input)
    output_path = Path(args.output)
    if not input_path.exists():
        raise SystemExit(f"Missing input workbook: {input_path}")

    wb = _load_workbook()(input_path)
    ws = wb[args.sheet] if args.sheet else wb.active
    headers = _ensure_headers(ws)
    _style_output_headers(ws, headers)

    processed = called = errors = 0
    for row in range(2, ws.max_row + 1):
        if args.limit is not None and processed >= args.limit:
            break
        title = _row_text(ws, row, headers, "title")
        description = _row_text(ws, row, headers, "description")
        if not title and not description:
            continue
        processed += 1

        if not args.overwrite and str(ws.cell(row=row, column=headers["llm_family"]).value or "").strip():
            continue

        posting = {
            "job_id": _row_text(ws, row, headers, "job_id"),
            "company_name": _row_text(ws, row, headers, "company_name"),
            "title": title,
            "description": description,
            "skills_desc": _row_text(ws, row, headers, "skills_desc"),
        }
        if args.dry_run:
            print(f"\n--- dry-run row {row} ---")
            print(prompt_for_posting(posting)[:2000])
            continue

        called += 1
        try:
            result = classify_posting(posting, api_key=api_key, model=model)
            ws.cell(row=row, column=headers["llm_family"], value=result["family"])
            ws.cell(row=row, column=headers["llm_confidence"], value=result["confidence"])
            ws.cell(row=row, column=headers["llm_rationale"], value=result["rationale"])
            ws.cell(row=row, column=headers["llm_error"], value="")
        except Exception as exc:  # keep row-level failures auditable in the workbook
            errors += 1
            ws.cell(row=row, column=headers["llm_family"], value="")
            ws.cell(row=row, column=headers["llm_confidence"], value="")
            ws.cell(row=row, column=headers["llm_rationale"], value="")
            ws.cell(row=row, column=headers["llm_error"], value=str(exc)[:1000])

        if args.delay > 0:
            time.sleep(args.delay)

    if args.dry_run:
        print(f"\nDry run completed for {processed} row(s). No output workbook written.")
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(
        f"Wrote {output_path} "
        f"(processed={processed}, api_calls={called}, errors={errors}, model={model})"
    )


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--input", default=str(ART / "sample_postings.xlsx"), help="Input .xlsx path.")
    p.add_argument("--output", default=str(ART / "sample_postings_with_llm_family.xlsx"), help="Output .xlsx path.")
    p.add_argument("--sheet", default="", help="Sheet name. Defaults to the active sheet.")
    p.add_argument("--limit", type=int, default=None, help="Limit rows processed.")
    p.add_argument("--delay", type=float, default=0.0, help="Seconds to sleep between API calls.")
    p.add_argument("--overwrite", action="store_true", help="Reclassify rows with llm_family.")
    p.add_argument("--dry-run", action="store_true", help="Print prompts without calling the API.")
    return p.parse_args()


if __name__ == "__main__":
    classify_workbook(parse_args())
